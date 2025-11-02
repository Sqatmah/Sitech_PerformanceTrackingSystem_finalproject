from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import timedelta
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import csv
from io import StringIO

from accounts.models import User, StudentProfile
from assignments.models import Assignment, Submission, Grade
from .models import Notification


@login_required
def home(request):
    """Main dashboard view - role-based content"""
    user = request.user
    context = {
        'user': user,
        'notifications': user.notifications.filter(is_read=False)[:5],
    }
    
    if user.is_admin:
        # Admin dashboard
        context.update({
            'total_users': User.objects.count(),
            'total_students': User.objects.filter(role='student').count(),
            'total_managers': User.objects.filter(role='manager').count(),
            'total_assignments': Assignment.objects.count(),
            'active_assignments': Assignment.objects.filter(is_active=True).count(),
            'total_submissions': Submission.objects.count(),
            'pending_submissions': Submission.objects.filter(status='submitted').count(),
            'recent_submissions': Submission.objects.select_related(
                'assignment', 'student'
            ).order_by('-submitted_at')[:10],
        })
        
    elif user.is_manager:
        # Manager dashboard
        my_students = User.objects.filter(student_profile__manager=user)
        my_assignments = Assignment.objects.filter(created_by=user)
        
        context.update({
            'my_students_count': my_students.count(),
            'my_assignments_count': my_assignments.count(),
            'active_assignments_count': my_assignments.filter(is_active=True).count(),
            'pending_submissions': Submission.objects.filter(
                assignment__created_by=user,
                status='submitted'
            ).count(),
            'recent_submissions': Submission.objects.filter(
                assignment__created_by=user
            ).select_related('assignment', 'student').order_by('-submitted_at')[:10],
            'my_students': my_students[:10],
        })
        
    elif user.is_student:
        # Student dashboard
        my_submissions = Submission.objects.filter(student=user)
        pending_assignments = Assignment.objects.filter(
            assigned_to=user,
            is_active=True
        ).exclude(
            submissions__student=user
        )
        
        context.update({
            'total_assignments': Assignment.objects.filter(assigned_to=user).count(),
            'completed_assignments': my_submissions.filter(status='graded').count(),
            'pending_assignments': pending_assignments.count(),
            'average_grade': my_submissions.filter(
                grade__isnull=False
            ).aggregate(avg=Avg('grade__score'))['avg'] or 0,
            'recent_submissions': my_submissions.select_related(
                'assignment'
            ).order_by('-submitted_at')[:5],
            'upcoming_assignments': Assignment.objects.filter(
                assigned_to=user,
                is_active=True,
                due_date__gte=timezone.now()
            ).order_by('due_date')[:5],
        })
    
    return render(request, 'dashboard/home.html', context)


@login_required
def notifications(request):
    """View all notifications"""
    user_notifications = request.user.notifications.all()
    
    # Mark all as read when viewing
    user_notifications.filter(is_read=False).update(is_read=True)
    
    return render(request, 'dashboard/notifications.html', {
        'notifications': user_notifications
    })


@login_required
def mark_notification_read(request, notification_id):
    """Mark a specific notification as read"""
    notification = get_object_or_404(
        Notification, 
        id=notification_id, 
        recipient=request.user
    )
    notification.is_read = True
    notification.save()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    return redirect('dashboard:notifications')


@login_required
def dashboard_stats(request):
    """API endpoint for dashboard statistics (for charts)"""
    user = request.user
    
    if user.is_admin:
        # Admin stats
        stats = {
            'user_registrations': list(
                User.objects.extra(
                    select={'month': "strftime('%Y-%m', date_joined)"}
                ).values('month').annotate(count=Count('id')).order_by('month')
            ),
            'assignment_submissions': list(
                Submission.objects.extra(
                    select={'month': "strftime('%Y-%m', submitted_at)"}
                ).values('month').annotate(count=Count('id')).order_by('month')
            ),
            'grade_distribution': [],  # Will be calculated in template or with raw SQL
        }
        
    elif user.is_manager:
        # Manager stats
        stats = {
            'my_assignments_submissions': list(
                Submission.objects.filter(
                    assignment__created_by=user
                ).extra(
                    select={'month': "strftime('%Y-%m', submitted_at)"}
                ).values('month').annotate(count=Count('id')).order_by('month')
            ),
            'student_performance': list(
                Grade.objects.filter(
                    submission__assignment__created_by=user
                ).values('submission__student__username').annotate(
                    avg_score=Avg('score')
                ).order_by('-avg_score')[:10]
            ),
        }
        
    elif user.is_student:
        # Student stats
        stats = {
            'my_grades': list(
                Grade.objects.filter(
                    submission__student=user
                ).values('submission__assignment__title', 'score', 'letter_grade')
                .order_by('submission__assignment__due_date')
            ),
            'submission_timeline': list(
                Submission.objects.filter(
                    student=user
                ).extra(
                    select={'month': "strftime('%Y-%m', submitted_at)"}
                ).values('month').annotate(count=Count('id')).order_by('month')
            ),
        }
    
    return JsonResponse(stats)


@login_required
def export_students(request):
    """Export students data to Excel (admin/manager only)"""
    if not (request.user.is_admin or request.user.is_manager):
        messages.error(request, 'You do not have permission to export this data.')
        return redirect('dashboard:home')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Headers
    headers = ['Username', 'First Name', 'Last Name', 'Email', 'Student ID', 
               'Manager', 'Enrollment Date', 'Status']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Get students based on user role
    if request.user.is_admin:
        students = User.objects.filter(role='student').select_related('student_profile')
    else:  # manager
        students = User.objects.filter(
            role='student',
            student_profile__manager=request.user
        ).select_related('student_profile')
    
    # Add data
    for student in students:
        profile = getattr(student, 'student_profile', None)
        ws.append([
            student.username,
            student.first_name,
            student.last_name,
            student.email,
            profile.student_id if profile else '',
            profile.manager.get_full_name() if profile and profile.manager else '',
            profile.enrollment_date.strftime('%Y-%m-%d') if profile and profile.enrollment_date else '',
            'Active' if profile and profile.is_active else 'Inactive'
        ])
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    wb.save(response)
    
    return response


@login_required
def export_assignments(request):
    """Export assignments data to CSV (admin/manager only)"""
    if not (request.user.is_admin or request.user.is_manager):
        messages.error(request, 'You do not have permission to export this data.')
        return redirect('dashboard:home')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=assignments.csv'
    
    writer = csv.writer(response)
    writer.writerow(['Title', 'Created By', 'Due Date', 'Priority', 'Status', 
                    'Assigned Students', 'Submissions', 'Average Grade'])
    
    # Get assignments based on user role
    if request.user.is_admin:
        assignments = Assignment.objects.all()
    else:  # manager
        assignments = Assignment.objects.filter(created_by=request.user)
    
    for assignment in assignments:
        avg_grade = assignment.submissions.filter(
            grade__isnull=False
        ).aggregate(avg=Avg('grade__score'))['avg']
        
        writer.writerow([
            assignment.title,
            assignment.created_by.get_full_name(),
            assignment.due_date.strftime('%Y-%m-%d %H:%M') if assignment.due_date else '',
            assignment.get_priority_display(),
            'Active' if assignment.is_active else 'Inactive',
            assignment.assigned_to.count(),
            assignment.submissions.count(),
            f"{avg_grade:.2f}" if avg_grade else 'N/A'
        ])
    
    return response
